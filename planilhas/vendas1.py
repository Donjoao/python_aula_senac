from openpyxl import Workbook
from statistics import mean
# Utilziei o codigo pip install openpyxl no terminal para instalar o openpyxl.
# Utilizei o codigo pip install pandas no terminal para instalar o pandas.
wb = Workbook()
ws = wb.active

ws['A1'] = "Dia"
ws['B1'] = "Vendedor"
ws['C1'] = "Produtos"
ws['D1'] = "Unidades"
ws['E1'] = "Preço"

#Dia
ws['A2'] = 2
ws['A3'] = 2
ws['A4'] = 3
ws['A5'] = 3
ws['A6'] = 5

#Vendedor
ws['B2'] = "Jorge"
ws['B3'] = "Marilia"
ws['B4'] = "Sofia"
ws['B5'] = "João"
ws['B6'] = "Jess"

#Produtos
ws['C2'] = "Celular"
ws['C3'] = "Carregador"
ws['C4'] = "Cabo USB"
ws['C5'] = "Cabo HDMI"
ws['C6'] = "Fone P2"

#Unidades
ws['D2'] = "2"
ws['D3'] = "1"
ws['D4'] = "1"
ws['D5'] = "1"
ws['D6'] = "4"

#Preços
ws['B2'] = 890
ws['B3'] = 85
ws['B4'] = 25
ws['B5'] = 45
ws['B6'] = 15

wb.save('vendas1.xlsx')