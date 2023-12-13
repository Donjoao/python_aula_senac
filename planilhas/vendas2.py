from openpyxl import Workbook


# Utilziei o codigo pip install openpyxl no terminal para instalar o openpyxl.
# Utilizei o codigo pip install pandas no terminal para instalar o pandas.

wb = Workbook()
ws = wb.active

#Head
ws['A1'] = "Dia"
ws['B1'] = "Vendedor"
ws['C1'] = "Produtos"
ws['D1'] = "Unidades"
ws['E1'] = "Preço"

#Dia
ws['A2'] = 7
ws['A3'] = 7
ws['A4'] = 8
ws['A5'] = 9
ws['A6'] = 11

#Vendedor
ws['B2'] = "Jorge"
ws['B3'] = "Marilia"
ws['B4'] = "Sofia"
ws['B5'] = "João"
ws['B6'] = "Jess"

#Produtos
ws['C2'] = "TV"
ws['C3'] = "TV"
ws['C4'] = "PS2"
ws['C5'] = "Cabo HDMI"
ws['C6'] = "Fone P2"

#Unidades
ws['D2'] = "1"
ws['D3'] = "1"
ws['D4'] = "1"
ws['D5'] = "1"
ws['D6'] = "4"

#Preços
ws['E2'] = 1400
ws['E3'] = 1400
ws['E4'] = 350
ws['E5'] = 45
ws['E6'] = 15

wb.save('vendas2.xlsx')