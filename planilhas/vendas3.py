from openpyxl import Workbook


# Utilziei o codigo pip install openpyxl no terminal para instalar o openpyxl.
# Utilizei o codigo pip install pandas no terminal para instalar o pandas.

wb = Workbook()
ws = wb.active

#Head
ws['A1'] = "Dia"
ws['B1'] = "Vendedor"
ws['C1'] = "Produtos"
ws['D1'] = "Unidades"
ws['E1'] = "Preço"

#Dia
ws['A2'] = 12
ws['A3'] = 12


#Vendedor
ws['B2'] = "Jorge"
ws['B3'] = "Marilia"


#Produtos
ws['C2'] = "PS2"
ws['C3'] = "Carregador"


#Unidades
ws['D2'] = 1
ws['D3'] = 2


#Preços
ws['E2'] = 350
ws['E3'] = 85

wb.save('vendas3.xlsx')