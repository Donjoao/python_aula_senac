from openpyxl import Workbook
from statistics import mean
# Utilziei o codigo pip install openpyxl no terminal para instalar o openpyxl.
# Utilizei o codigo pip install pandas no terminal para instalar o pandas.
wb = Workbook()
ws = wb.active

ws['A1'] = "Nome"
ws['B1'] = "Produto"
ws['C1'] = "Preço"
ws['A2'] = "Jose"
ws['B2'] = "Celular"
ws['C2'] = 1000
ws['A3'] = "Maria"
ws['B3'] = "Celular"
ws['C3'] = 980
ws['B5'] = "Total"
ws['C5'] = sum([ws["C2"].value,ws["C3"].value])
ws['C6'] = min([ws["C2"].value,ws["C3"].value])
ws['C7'] = mean([ws["C2"].value, ws["C3"].value])
wb.save('my_excel_file.xlsx')

